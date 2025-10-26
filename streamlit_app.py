import io
import os
import zipfile
from datetime import datetime

import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import WriteOnlyCell


st.set_page_config(page_title="Excel Shrinker", page_icon="🧹", layout="centered")
st.title("🧹 Excel Shrinker — Kecilkan ukuran tanpa mengubah isi")
st.caption("Pilih metode kompresi sesuai kebutuhanmu. Mode Lossless aman untuk semua file .xlsx/.xlsm.")


# ===== Utilities =====
def human_size(num_bytes: int) -> str:
    for unit in ["B", "KB", "MB", "GB"]:
        if num_bytes < 1024.0:
            return f"{num_bytes:,.2f} {unit}"
        num_bytes /= 1024.0
    return f"{num_bytes:,.2f} TB"


def detect_vba(xlsx_bytes: bytes) -> bool:
    try:
        with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as z:
            return any(name.lower().startswith("xl/") and name.lower().endswith("vbaProject.bin".lower())
                       for name in z.namelist())
    except zipfile.BadZipFile:
        return False


def recompress_xlsx(xbytes: bytes, compresslevel: int = 9) -> bytes:
    """Lossless: Re-zip semua entry. Tidak mengubah isi/struktur workbook."""
    src_io = io.BytesIO(xbytes)
    out_io = io.BytesIO()
    with zipfile.ZipFile(src_io, "r") as zin, \
         zipfile.ZipFile(out_io, "w") as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            # File gambar/binary biasanya tidak dapat diperkecil dengan deflate — simpan apa adanya.
            is_binary = info.filename.lower().endswith((
                ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".emf", ".wmf", ".bin"
            ))
            if is_binary:
                zout.writestr(info.filename, data, compress_type=zipfile.ZIP_STORED)
            else:
                # compresslevel pada writestr tersedia di Python 3.7+
                zout.writestr(
                    info.filename,
                    data,
                    compress_type=zipfile.ZIP_DEFLATED,
                    compresslevel=compresslevel
                )
    out_io.seek(0)
    return out_io.getvalue()


def minify_xlsx(xbytes: bytes, keep_number_formats: bool = True) -> bytes:
    """
    Copy nilai & formula saja ke workbook baru (write_only) -> buang styling, objek, gambar, dsb.
    Opsi: jaga number format agar tampilan angka/tanggal tetap nyaman dibaca.
    """
    src = io.BytesIO(xbytes)
    wb_src = load_workbook(src, data_only=False, keep_links=True)  # formulas tetap string "=..."
    # IMPORTANT: openpyxl write_only tidak mendukung VBA; fungsi ini hanya untuk non-macro workbook.
    wb_dst = Workbook(write_only=True)

    # Gunakan sheet pertama default pada wb_dst
    first = True
    for ws in wb_src.worksheets:
        if first:
            ws_new = wb_dst.active
            ws_new.title = ws.title
            first = False
        else:
            ws_new = wb_dst.create_sheet(title=ws.title)

        # Pertahankan status sheet (visible/hidden) — ini tidak mengubah isi data.
        try:
            ws_new.sheet_state = ws.sheet_state
        except Exception:
            pass

        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            out_row = []
            for c in row:
                val = c.value  # nilai atau string formula seperti "=SUM(A1:A3)"
                wcell = WriteOnlyCell(ws_new, value=val)

                # Pertahankan hyperlink (bagian dari isi semantik)
                if getattr(c, "hyperlink", None):
                    link = getattr(c.hyperlink, "target", None) or str(c.hyperlink)
                    if link:
                        wcell.hyperlink = link

                if keep_number_formats:
                    # Menjaga number format agar tanggal/angka tetap terbaca. (Tidak menyalin styling lain)
                    try:
                        wcell.number_format = c.number_format
                    except Exception:
                        pass

                out_row.append(wcell)
            ws_new.append(out_row)

    out = io.BytesIO()
    wb_dst.save(out)
    out.seek(0)
    return out.getvalue()


# ===== UI =====
uploaded = st.file_uploader(
    "Unggah file Excel (.xlsx atau .xlsm)", type=["xlsx", "xlsm"], accept_multiple_files=False
)

mode = st.radio(
    "Pilih metode kompresi:",
    options=["Lossless (Recompress ZIP)", "Minify (Value & Formula Only)"],
    help=(
        "• Lossless: Tidak mengubah isi/format/objek sama sekali; hanya kompres ulang ZIP di dalam .xlsx/.xlsm.\n"
        "• Minify: Salin nilai & formula ke workbook baru (tanpa styling/objek) agar lebih kecil."
    ),
)

keep_formats = st.checkbox(
    "Pertahankan format angka/tanggal (disarankan untuk Minify)",
    value=True,
)

compress_lvl = st.slider(
    "Tingkat kompresi (Lossless)", min_value=1, max_value=9, value=9,
    help="Hanya berlaku untuk mode Lossless."
)

if uploaded is not None:
    name, ext = os.path.splitext(uploaded.name)
    ext = ext.lower()
    in_bytes = uploaded.read()
    in_size = len(in_bytes)
    st.write(f"**Ukuran asli:** {human_size(in_size)}")

    has_macro = detect_vba(in_bytes) if ext == ".xlsm" else False
    if mode.startswith("Minify") and has_macro:
        st.warning("File ini berisi **macro (VBA)**. Mode **Minify** tidak mendukung macro. "
                   "Gunakan mode **Lossless** agar isi tetap utuh.")
    else:
        if st.button("Proses sekarang", type="primary"):
            try:
                if mode.startswith("Lossless"):
                    out_bytes = recompress_xlsx(in_bytes, compresslevel=int(compress_lvl))
                    out_ext = ext  # pertahankan .xlsx atau .xlsm
                else:
                    out_bytes = minify_xlsx(in_bytes, keep_number_formats=bool(keep_formats))
                    out_ext = ".xlsx"  # workbook baru tanpa macro

                out_size = len(out_bytes)
                delta = in_size - out_size
                pct = (delta / in_size * 100.0) if in_size else 0.0

                st.success(
                    f"Selesai! Ukuran baru: **{human_size(out_size)}** "
                    f"({human_size(delta)} lebih kecil, ~{pct:.2f}%)."
                )

                ts = datetime.now().strftime("%Y%m%d-%H%M%S")
                out_name = f"{name}.shrink-{ts}{out_ext}"
                st.download_button(
                    "💾 Unduh hasil",
                    data=out_bytes,
                    file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    if out_ext == ".xlsx"
                    else "application/vnd.ms-excel.sheet.macroEnabled.12",
                )

                with st.expander("Catatan Teknis"):
                    st.markdown(
                        """
                        - **Lossless** hanya melakukan re-konstruksi arsip ZIP internal dengan kompresi DEFLATE level tinggi.
                          Tidak mengubah XML/relasi/part apa pun.
                        - **Minify** menulis ulang workbook secara `write_only`:
                          menyalin **nilai & formula** persis, menghapus styling/objek agar ukuran jauh lebih kecil.
                          Opsi **format angka/tanggal** menjaga tampilan angka/tanggal tetap nyaman.
                        - Elemen yang memang bukan "isi" (chart, gambar, pivot cache, conditional formatting, dsb.)
                          memang dihapus pada mode Minify untuk penghematan maksimal.
                        """
                    )
            except zipfile.BadZipFile:
                st.error("Berkas tidak valid/korup atau bukan file .xlsx/.xlsm.")
            except Exception as e:
                st.error(f"Terjadi kesalahan: {e}")
else:
    st.info("Unggah file .xlsx atau .xlsm untuk mulai mengecilkan ukuran.")
